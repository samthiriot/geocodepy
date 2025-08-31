import os
import sys

from shapely import LineString
from shapely.geometry import Point
import pandas

from geocodepy.geocoders import IGNFrance
from geocodepy.geocoders import Nominatim

from geocodepy.geocoders.banfrance import BANFrance


import logging

logging.basicConfig(level=logging.DEBUG)

# from http.client import HTTPConnection
# HTTPConnection.debuglevel = 1

sample_size = 10000
filter_commune = None  # "FONTAINEBLEAU"

res_directory = "/tmp/"
res_file_excel = "/tmp/dle_gaz_2023_geocoded.xlsx"

print("reading data...")
df = pandas.read_csv("examples/data/dle_gaz_2023.csv",
                     sep=";",
                     #nrows=10,
                     header=0,
                     skiprows=1,  # ignore the first row of header
                     usecols=["ADRESSE", "NOM_COMMUNE"]
                     )

if filter_commune is not None:
    print("filtering on", filter_commune)
    df = df[df["NOM_COMMUNE"] == filter_commune]

if sample_size is not None and sample_size < len(df):
    print("sampling", sample_size, "addresses") 
    df = df.sample(n=sample_size, random_state=42)

print(df)
df["query"] = df["ADRESSE"] + ", " + df["NOM_COMMUNE"]


def geocode_dataframe(df, geocoder):

    def progress_callback(i, total):
        print(".", end='')
        sys.stdout.flush()
        
    results = geocoder.geocode_batch(df["query"], timeout=10)
    print()

    df["label"] = [result.address if result is not None else None
                   for result in results]
    df["latitude"] = [float(result.latitude) if result is not None else None
                      for result in results]
    df["longitude"] = [float(result.longitude) if result is not None else None
                       for result in results]

    decoded_raw = pandas.DataFrame.from_records([
        {key: value for key, value in result.raw.items() if key not in ["properties"]}
        if result is not None else {}
        for result in results])
    decoded_raw.drop(['label', 'latitude', 'longitude'], axis=1, inplace=True, errors='ignore')

    has_properties = False
    for result in results:
        if result is None or result.raw is None:
            continue
        has_properties = "properties" in result.raw
        break

    df.reset_index(drop=True, inplace=True)
    if has_properties:
        decoded_properties = pandas.DataFrame.from_records([
            result.raw["properties"]
            if result is not None else {}
            for result in results])
        decoded_properties.drop(['label', 'latitude', 'longitude'], axis=1, inplace=True, errors='ignore')
        result = pandas.concat([df, decoded_raw, decoded_properties],
                               axis=1, join='outer')
    else:
        result = pandas.concat([df, decoded_raw],
                               axis=1, join='outer')

    # ensure the type of a few columns
    for column in ['score', 'lat', 'lon', 'importance', 'x', 'y', 'score_next']:
        if column in result.columns:
            result[column] = result[column].astype(float)

    return result


with pandas.ExcelWriter(res_file_excel) as writer:

    print("exporting sample...")
    df.to_excel(writer, sheet_name="sample")

    dataframes = {}

    print("geocoding with BAN...")
    results = geocode_dataframe(df, BANFrance())
    print("exporting BAN results...")
    results.to_excel(writer, sheet_name="BAN")
    dataframes["BAN"] = results

    print("geocoding with IGN...")
    results = geocode_dataframe(df, IGNFrance())
    print("exporting IGN results...")
    results.to_excel(writer, sheet_name="IGN")
    dataframes["IGN"] = results

    print("geocoding with Nominatim...")
    results = geocode_dataframe(df, Nominatim(user_agent="geocodepy-test-batch"))
    print("exporting Nominatim results...")
    results.to_excel(writer, sheet_name="Nominatim")
    dataframes["Nominatim"] = results

# format the Excel
import openpyxl as px
from openpyxl.utils import get_column_letter

wb = px.load_workbook(res_file_excel)
for ws in wb.worksheets:
    # activate autofilter
    ws.auto_filter.ref = ws.dimensions
    # try to autofit
    for i in range(1, ws.max_column + 1):
        ws.column_dimensions[get_column_letter(i)].bestFit = True
        ws.column_dimensions[get_column_letter(i)].auto_size = True

wb.save(res_file_excel)

# print("exported in Excel", res_file_excel)

# merged = None
for geocoder, dataframe in dataframes.items():
    dataframe["geocoder"] = geocoder
    #dataframe = dataframe.reset_index(drop=True)
    dataframe.columns = pandas.io.common.dedup_names(dataframe.columns, is_potential_multiindex=False)


#     if merged is None:
#         merged = dataframe
#     else:
#         pandas.merge(merged, dataframe, on="query", how="outer")

concatenated = pandas.concat(list(dataframes.values()), ignore_index=True, axis=0, join='outer')


import geopandas
# export as GeoJSON
gdf = geopandas.GeoDataFrame(
    concatenated, geometry=geopandas.points_from_xy(concatenated.longitude, concatenated.latitude), crs="EPSG:4326"
)
gdf.to_file(os.path.join(res_directory, "dle_gaz_2023_geocoded.geojson"), driver="GeoJSON")


def create_layer_lines_difference(df1, df2):
    juxtaposed = pandas.concat([df1, df2], axis=1, join='outer')
    juxtaposed.columns = pandas.io.common.dedup_names(juxtaposed.columns, is_potential_multiindex=False)
    #print(juxtaposed.head())

    gdf = geopandas.GeoDataFrame(
        juxtaposed,
        geometry=[LineString([Point(x1, y1), Point(x2, y2)])
                  for x1, y1, x2, y2 in zip(df1.longitude, df1.latitude, df2.longitude, df2.latitude)
                  ],
        crs="EPSG:4326"
    )
    print(juxtaposed.head())
    return gdf


create_layer_lines_difference(dataframes["IGN"], dataframes["BAN"]).to_file(os.path.join(res_directory, "ign2ban.geojson"), driver="GeoJSON")
create_layer_lines_difference(dataframes["IGN"], dataframes["Nominatim"]).to_file(os.path.join(res_directory, "ign2nominatim.geojson"), driver="GeoJSON")

